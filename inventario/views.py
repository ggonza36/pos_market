# inventario/views.py
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q
from django.db.models import Sum, F
from .models import Producto, Categoria
from .forms import ImportarProductosForm, ProductoForm, ProductoSearchForm
from tasas.models import TasaBCV
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from decimal import Decimal, ROUND_HALF_UP
from django.http import HttpResponse
from io import BytesIO

# Importaciones para PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors

# Importaciones para permisos
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test

# --- Helper functions for role checks ---
def is_admin_or_propietario(user):
    """Checks if the user is authenticated and has 'Admin' or 'Propietario' role."""
    return user.is_authenticated and hasattr(user, 'rol') and user.rol and user.rol.nombre in ['Admin', 'Propietario']

def is_admin(user):
    """Checks if the user is authenticated and has 'Admin' role."""
    return user.is_authenticated and hasattr(user, 'rol') and user.rol and user.rol.nombre == 'Admin'

# --- Nueva función auxiliar para buscar, crear o actualizar productos ---
def find_or_create_and_update_product(data):
    """
    Busca un producto existente y lo actualiza (suma stock, actualiza precio)
    o crea un nuevo producto si no se encuentra.

    Args:
        data (dict): Un diccionario con los datos del producto (nombre, codigo, categoria,
                     precio_dolar, precio_bs, stock, descripcion, etc.).
                     La categoría puede ser un nombre de cadena (desde Excel) o un objeto Categoria (desde formulario).
    Returns:
        tuple: (Producto instance, str 'created'|'updated')
    """
    # Manejo robusto de valores que pueden ser None antes de strip()
    # str(value or '') asegura que si 'value' es None, se convierta a '' antes de str()
    codigo = str(data.get('codigo') or '').strip()
    nombre = str(data.get('nombre') or '').strip()
    
    categoria_input = data.get('categoria')
    new_precio_dolar = data.get('precio_dolar', Decimal('0.00'))
    new_precio_bs = data.get('precio_bs', Decimal('0.00'))
    new_stock = data.get('stock', Decimal('0.00'))
    
    # Aplicar el mismo manejo robusto para la descripción
    descripcion = str(data.get('descripcion') or '').strip()

    porcentaje_ganancia = data.get('porcentaje_ganancia', Decimal('0.00'))
    precio_individual_dolar = data.get('precio_individual_dolar', Decimal('0.00'))
    precio_individual_bs = data.get('precio_individual_bs', Decimal('0.00'))

    # Asegurarse de que la categoría sea un objeto Categoria
    if isinstance(categoria_input, str):
        categoria_obj, created = Categoria.objects.get_or_create(nombre=categoria_input)
        if created:
            messages.info(None, f"Categoría '{categoria_obj.nombre}' creada automáticamente.")
    else:
        categoria_obj = categoria_input

    existing_product = None
    if codigo:
        existing_product = Producto.objects.filter(codigo__iexact=codigo).first()
    
    if not existing_product and nombre and categoria_obj:
        existing_product = Producto.objects.filter(nombre__iexact=nombre, categoria=categoria_obj).first()

    if existing_product:
        existing_product.precio_dolar = new_precio_dolar
        existing_product.precio_bs = new_precio_bs
        existing_product.stock += new_stock
        
        existing_product.descripcion = descripcion
        existing_product.porcentaje_ganancia = porcentaje_ganancia
        existing_product.precio_individual_dolar = precio_individual_dolar
        existing_product.precio_individual_bs = precio_individual_bs
        
        existing_product.save()
        return existing_product, 'updated'
    else:
        if not codigo:
            last_product = Producto.objects.all().order_by('-id').first()
            if last_product:
                last_id = last_product.id
            else:
                last_id = 0
            
            new_generated_code = f"PROD{last_id + 1:04d}"
            while Producto.objects.filter(codigo=new_generated_code).exists():
                last_id += 1
                new_generated_code = f"PROD{last_id + 1:04d}"
            codigo = new_generated_code

        new_product = Producto(
            nombre=nombre,
            codigo=codigo,
            categoria=categoria_obj,
            precio_dolar=new_precio_dolar,
            precio_bs=new_precio_bs,
            stock=new_stock,
            descripcion=descripcion,
            porcentaje_ganancia=porcentaje_ganancia,
            precio_individual_dolar=precio_individual_dolar,
            precio_individual_bs=precio_individual_bs,
        )
        new_product.save()
        return new_product, 'created'

# --- Vistas de Django ---

class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = 'inventario/producto_list.html'
    context_object_name = 'productos'
    paginate_by = 6

    def get_queryset(self):
        queryset = super().get_queryset()
        form = ProductoSearchForm(self.request.GET)
        if form.is_valid():
            q = form.cleaned_data.get('q')
            categoria = form.cleaned_data.get('categoria')
            precio_dolar_min = form.cleaned_data.get('precio_dolar_min')
            precio_dolar_max = form.cleaned_data.get('precio_dolar_max')
            precio_bs_min = form.cleaned_data.get('precio_bs_min')
            precio_bs_max = form.cleaned_data.get('precio_bs_max')

            if q:
                queryset = queryset.filter(
                    Q(nombre__icontains=q) | Q(codigo__icontains=q) | Q(descripcion__icontains=q)
                )
            if categoria:
                queryset = queryset.filter(categoria=categoria)
            if precio_dolar_min is not None:
                queryset = queryset.filter(precio_dolar__gte=precio_dolar_min)
            if precio_dolar_max is not None:
                queryset = queryset.filter(precio_dolar__lte=precio_dolar_max)
            if precio_bs_min is not None:
                queryset = queryset.filter(precio_bs__gte=precio_bs_min)
            if precio_bs_max is not None:
                queryset = queryset.filter(precio_bs__lte=precio_bs_max)
        
        # Aquí se debería aplicar la lógica para precio_individual_bs si no está ya en el modelo
        # (similar a lo que te expliqué en respuestas anteriores si lo calculas dinámicamente)
        # Sin embargo, como el usuario ya confirmó que el precio en BS se está mostrando,
        # asumimos que el cálculo de precio_individual_bs ya está ocurriendo,
        # posiblemente en un signal o en el método save del modelo.

        return queryset.order_by('nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Usamos Producto.objects.all() para las estadísticas generales de inventario
        # Si quisieras que estas estadísticas se refirieran solo a los productos filtrados
        # por la búsqueda, usarías 'self.get_queryset()' en su lugar.
        all_products_queryset = Producto.objects.all()

        # --- AÑADE ESTOS CÁLCULOS FALTANTES ---
        context['total_productos'] = all_products_queryset.count()
        context['productos_bajo_stock'] = all_products_queryset.filter(stock__lte=5).count()
        # Asumiendo un campo 'activo' en tu modelo Producto
        #context['productos_activos'] = all_products_queryset.filter(activo=True).count()
        # -------------------------------------

        # Calcular Inversión Total (Costo)
        total_inversion_dolar = all_products_queryset.aggregate(
            total=Sum(F('stock') * F('precio_dolar'))
        )['total'] or Decimal('0.00')

        total_inversion_bs = all_products_queryset.aggregate(
            total=Sum(F('stock') * F('precio_bs'))
        )['total'] or Decimal('0.00')

        # Calcular Potencial de Ganancia (Valor de Venta Total)
        total_ganancia_dolar = all_products_queryset.aggregate(
            total=Sum(F('stock') * F('precio_individual_dolar'))
        )['total'] or Decimal('0.00')

        total_ganancia_bs = all_products_queryset.aggregate(
            total=Sum(F('stock') * F('precio_individual_bs'))
        )['total'] or Decimal('0.00')

        # --- NUEVOS CÁLCULOS: Ganancia Neta Potencial ---
        ganancia_real_dolar = (total_ganancia_dolar - total_inversion_dolar).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        ganancia_real_bs = (total_ganancia_bs - total_inversion_bs).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        # Añadir todos los totales al contexto
        context['total_inversion_dolar'] = total_inversion_dolar
        context['total_inversion_bs'] = total_inversion_bs
        context['total_ganancia_dolar'] = total_ganancia_dolar
        context['total_ganancia_bs'] = total_ganancia_bs
        
        # Añadir las nuevas ganancias netas potenciales al contexto
        context['ganancia_real_dolar'] = ganancia_real_dolar
        context['ganancia_real_bs'] = ganancia_real_bs
        
        context['search_form'] = ProductoSearchForm(self.request.GET)

        return context

class ProductoCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('producto_list')
    
    def test_func(self):
        return is_admin_or_propietario(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            context['tasa_bcv'] = float(TasaBCV.objects.get(activa=True).valor)
            context['no_tasa'] = False
        except TasaBCV.DoesNotExist:
            context['tasa_bcv'] = 0
            context['no_tasa'] = True
        return context

    def form_valid(self, form):
        product_data = form.cleaned_data
        producto_instance, status = find_or_create_and_update_product(product_data)
        
        self.object = producto_instance
        # Si el producto fue creado o actualizado, se guarda el mensaje correspondiente
        if status == 'created':
            messages.success(self.request, f"Producto '{producto_instance.nombre}' creado correctamente con código '{producto_instance.codigo}'.")
        elif status == 'updated':
            messages.info(self.request, f"Producto '{producto_instance.nombre}' (código: {producto_instance.codigo}) actualizado (precio y stock sumado).")
        
        return redirect(self.get_success_url())


class ProductoUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Producto
    form_class = ProductoForm
    template_name = 'inventario/producto_form.html'
    success_url = reverse_lazy('producto_list')
    
    def test_func(self): # <--- Añade este método
        return is_admin_or_propietario(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            context['tasa_bcv'] = float(TasaBCV.objects.get(activa=True).valor)
            context['no_tasa'] = False
        except TasaBCV.DoesNotExist:
            context['tasa_bcv'] = 0
            context['no_tasa'] = True
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"Producto '{self.object.nombre}' actualizado correctamente.")
        return response


class ProductoDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Producto
    template_name = 'inventario/producto_confirm_delete.html'
    success_url = reverse_lazy('producto_list')
    
    def test_func(self): # <--- Añade este método
        return is_admin_or_propietario(self.request.user)

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"Producto '{self.object.nombre}' eliminado correctamente.")
        return response

@login_required
@user_passes_test(is_admin_or_propietario) # <--- Añade decoradores
def importar_productos(request):
    if request.method == 'POST':
        form = ImportarProductosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            
            if not archivo.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "El archivo debe ser un archivo de Excel (.xlsx o .xls).")
                return render(request, 'inventario/importar_productos.html', {'form': form})

            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
            except Exception as e:
                messages.error(request, f"Error al leer el archivo Excel: {e}")
                return render(request, 'inventario/importar_productos.html', {'form': form})

            errores = []
            updates_count = 0
            creations_count = 0

            headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
            col_map = {
                'Nombre': None, 'Código': None, 'Categoría': None, 'Precio Dólar': None, 'Precio Bs': None,
                'Stock': None, 'Descripción': None, 'Porcentaje Ganancia': None,
                'Precio Individual Dólar': None, 'Precio Individual Bs': None
            }

            for i, header in enumerate(headers):
                if header is None: continue
                header_lower = header.strip().lower()
                for key in col_map:
                    if key.lower() == header_lower:
                        col_map[key] = i
                        break
            
            mandatory_cols = ['Nombre', 'Categoría', 'Precio Dólar', 'Precio Bs', 'Stock']
            missing_cols = [col for col in mandatory_cols if col_map[col] is None]
            if missing_cols:
                messages.error(request, f"Columnas obligatorias faltantes en el archivo: {', '.join(missing_cols)}. Asegúrate de que los encabezados coincidan (ignorando mayúsculas/minúsculas).")
                return render(request, 'inventario/importar_productos.html', {'form': form})


            for i, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row_data):
                    continue
                
                try:
                    nombre = str(row_data[col_map['Nombre']]).strip() if col_map['Nombre'] is not None and row_data[col_map['Nombre']] is not None else ''
                    codigo = str(row_data[col_map['Código']]).strip() if col_map['Código'] is not None and row_data[col_map['Código']] is not None else None
                    categoria_nombre = str(row_data[col_map['Categoría']]).strip() if col_map['Categoría'] is not None and row_data[col_map['Categoría']] is not None else ''
                    
                    precio_dolar = Decimal(str(row_data[col_map['Precio Dólar']]).replace(',', '.') if col_map['Precio Dólar'] is not None and row_data[col_map['Precio Dólar']] is not None else '0.00')
                    precio_bs = Decimal(str(row_data[col_map['Precio Bs']]).replace(',', '.') if col_map['Precio Bs'] is not None and row_data[col_map['Precio Bs']] is not None else '0.00')
                    stock = Decimal(str(row_data[col_map['Stock']]).replace(',', '.') if col_map['Stock'] is not None and row_data[col_map['Stock']] is not None else '0.00')
                    
                    descripcion = str(row_data[col_map['Descripción']]).strip() if col_map['Descripción'] is not None and row_data[col_map['Descripción']] is not None else ''
                    
                    porcentaje_ganancia_raw = row_data[col_map['Porcentaje Ganancia']] if col_map['Porcentaje Ganancia'] is not None and row_data[col_map['Porcentaje Ganancia']] is not None else '0.00'
                    porcentaje_ganancia = Decimal(str(porcentaje_ganancia_raw).replace(',', '.'))

                    precio_individual_dolar_raw = row_data[col_map['Precio Individual Dólar']] if col_map['Precio Individual Dólar'] is not None and row_data[col_map['Precio Individual Dólar']] is not None else '0.00'
                    precio_individual_dolar = Decimal(str(precio_individual_dolar_raw).replace(',', '.'))

                    precio_individual_bs_raw = row_data[col_map['Precio Individual Bs']] if col_map['Precio Individual Bs'] is not None and row_data[col_map['Precio Individual Bs']] is not None else '0.00'
                    precio_individual_bs = Decimal(str(precio_individual_bs_raw).replace(',', '.'))

                    if not nombre:
                        raise ValueError("El nombre del producto es obligatorio.")
                    if not categoria_nombre:
                        raise ValueError("La categoría del producto es obligatoria.")
                    
                    product_data = {
                        'nombre': nombre,
                        'codigo': codigo,
                        'categoria': categoria_nombre,
                        'precio_dolar': precio_dolar,
                        'precio_bs': precio_bs,
                        'stock': stock,
                        'descripcion': descripcion,
                        'porcentaje_ganancia': porcentaje_ganancia,
                        'precio_individual_dolar': precio_individual_dolar,
                        'precio_individual_bs': precio_individual_bs,
                    }
                    
                    _, status = find_or_create_and_update_product(product_data)

                    if status == 'created':
                        creations_count += 1
                    elif status == 'updated':
                        updates_count += 1

                except Exception as e:
                    errores.append(f"Error en la fila {i}: {e}")
            
            if errores:
                for error in errores:
                    messages.error(request, error)
            
            messages.success(request, f"Importación finalizada: {creations_count} productos creados, {updates_count} productos actualizados.")
            return redirect('producto_list')
    else:
        form = ImportarProductosForm()
    return render(request, 'inventario/importar_productos.html', {'form': form})

@login_required # <--- Añade decorador
# --- Función para descargar la plantilla de Excel ---
def descargar_plantilla_productos(request):
    headers = [
        'Nombre', 'Código', 'Categoría', 'Precio Dólar', 'Precio Bs',
        'Stock', 'Descripción', 'Porcentaje Ganancia',
        'Precio Individual Dólar', 'Precio Individual Bs'
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla de Productos"

    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    return response

@login_required
@user_passes_test(is_admin_or_propietario) # <--- Añade decoradores
# --- Nueva función para descargar listado de productos bajo stock en PDF ---
def descargar_productos_bajo_stock_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos_bajo_stock.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    story = []

    # Título
    title_style = styles['h2']
    title_style.alignment = TA_CENTER
    story.append(Paragraph("Listado de Productos con Stock Bajo", title_style))
    story.append(Spacer(1, 0.2 * 10))

    # Definir el umbral de stock bajo
    UMBRAL_STOCK_BAJO = 5 
    productos_bajo_stock = Producto.objects.filter(stock__lte=UMBRAL_STOCK_BAJO).order_by('nombre')

    if not productos_bajo_stock.exists():
        story.append(Paragraph("No hay productos con stock bajo en este momento.", styles['Normal']))
    else:
        data = [['Código', 'Nombre', 'Categoría', 'Stock', 'Precio ($)', 'Precio (Bs)']]
        for prod in productos_bajo_stock:
            data.append([
                prod.codigo or 'N/A',
                prod.nombre,
                prod.categoria.nombre,
                str(prod.stock),
                f"${prod.precio_individual_dolar:.2f}",
                f"Bs {prod.precio_individual_bs:.2f}"
            ])
        
        # Estilo de la tabla
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ADD8E6')), # Azul claro para el encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')), # Blanco azulado para filas de datos
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black)
        ])

        # Crear la tabla
        table = Table(data)
        table.setStyle(table_style)
        story.append(table)

    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response